import re
from time import sleep
import openpyxl
#"متاسفانه سیستم با خطا مواجه شده است." in maliHTML.text


excell = openpyxl.load_workbook('coda.xlsx')
#sheet = excell.active

nameOfCompanies = []

def get_html(link):
    sleep(5)
    from selenium import webdriver
    from bs4 import BeautifulSoup
    url = link
    driver = webdriver.Chrome()
    driver.get(url)
    html_content = driver.page_source
    html_content = BeautifulSoup(html_content,'html.parser')
    driver.quit()
    sleep(5)
    return html_content

def get_exact_url(ctg):
    import urllib.parse

    base_url = "https://codal.ir/ReportList.aspx?search&Symbol="
    symbol = ctg

    # Encode the symbol using urllib.parse.quote()
    encoded_symbol = urllib.parse.quote(symbol, safe='')

    # Construct the complete URL
    complete_url = f"{base_url}{encoded_symbol}&LetterType=-1&AuditorRef=-1&PageNumber=1&Audited&NotAudited&IsNotAudited=false&Childs&Mains&Publisher=false&CompanyState=-1&Category=-1&CompanyType=-1&Consolidatable&NotConsolidatable"

    return complete_url


url = "https://codal.ir/ReportList.aspx?PageNumber="

url1 = "https://codal.ir/ReportList.aspx?search&LetterType=-1&AuditorRef=-1&PageNumber="
url2 = "&Audited&NotAudited&IsNotAudited=false&Childs&Mains&Publisher=false&CompanyState=-1&Category=1&CompanyType=-1&Consolidatable&NotConsolidatable"



for i in range(1,6495):
    link = url1 + str(i) + url2
    pageHTML = get_html(link)
    # find name of company
    allPosts = pageHTML.find_all('tr',attrs={'class':'table__row ng-scope'})

    for post in allPosts:
        
        name = post.find('td',attrs={'data-heading':'نماد'})    
        name = name.text.strip()  # اسم شرکت

        notif = post.find('a',attrs={'class':'letter-title ng-binding ng-scope'})
        matnNotif = notif.text.strip()   # عنوان اطلاعیه

        if (((name,matnNotif) not in nameOfCompanies) and ("صورت‌های مالی" in matnNotif)):
            
            nameOfCompanies.append((name,matnNotif))

            if ("تلفیقی" in matnNotif.split()):     # تلفیقی بودن یا نبودن
                 talfiqi = True
            else:
                 talfiqi = False

            

            date = "";dore = "";mtn = matnNotif.split();year=""   # تاریخ سالی و دوره چند ماهگی
            for d in range(len(mtn)):
                 if ("/" in mtn[d]):
                      date=mtn[d]
                      year = date.split("/")[0]
                 if ("ماهه" in mtn[d]):
                      dore = mtn[d-1]+" "+mtn[d]

            

            #date = date.split("/")[0]
            #nameOfCompanies.append((name,matnNotif))
            #linkOfCompany = get_exact_url(name)
            linkNotif = "https://www.codal.ir"+notif.get('href')
            sleep(3)
            #notifHTML = get_html(linkNotif)
            #companyPosts = companyHTML.find_all('tr',attrs={'class':'table__row ng-scope'})
            

            # صورت وضعیت مالی
            vaziatMali = []
            linkMali = linkNotif+"&sheetid=0"
            maliHTML = get_html(linkMali)
            maliRows = maliHTML.find('tbody').find_all('tr')
            if ("مشهود" in maliRows[2].text.split()):
                for j in range(2,len(maliRows)):
                    row = maliRows[j].text.split()
                    try:
                        if ((j==2) or (j==3) or (j==7) or (j==10) or (j==11) or (j==18) or (j==25) or (j==28) or (j==31) or (j==36) or (j==37) or (j==39) or (j==43) or (j==49) ):
                            #if ('۰' not in row):
                                var = row[0].strip()+" "+row[1]+" "+row[2]
                                val = row[3]
                                if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 
                                    #vaziatMali.append((var,val))

                        if ((j==4) or (j==48) or (j==5) or (j==6) or (j==8) or (j==14) or (j==15) or (j==19) or (j==24) or (j==26) or (j==27) or (j==35) or (j==42) or (j==44) or (j==50)):
                            #if ('۰' not in row):
                                var = row[0].strip()+" "+row[1]
                                val = row[2]
                                #val2 = row[-4] 
                                if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 

                        if ((j==12) or (j==23)):
                            #if ('۰' not in row):
                                    var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]
                                    val = row[4]
                                    if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 
                        
                        if ((j==13) or (j==17) or (j==29) or (j==41) or (j==51)): # or (j==38)
                            #if ('۰' not in row):
                                    var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]
                                    val = row[5]
                                    if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 
                        
                        if ((j==22) or (j==45) or (j==46)):
                            #if ('۰' not in row):
                                    var = row[0].strip()
                                    val = row[1]
                                    if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 
                    except:
                        vaziatMali = []
                         
                
                '''if (j==48):
                    #if ('۰' not in row):
                            var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]+" "+row[6]
                            val = row[7]
                            if (((var,val) not in row)):
                                vaziatMali.append((var,val)) '''
                
                #
            '''else:
                 if ("مشهود" in maliRows[3].text.split()):
                    for j in range(3,len(maliRows)):
                        row = maliRows[j].text.split()
                        if ((j==3) or (j==4) or (j==8) or (j==11) or (j==12) or (j==19) or (j==26) or (j==29) or (j==32) or (j==37) or (j==38) or (j==40) or (j==44) or (j==50) ):
                            #if ('۰' not in row):
                                var = row[0].strip()+" "+row[1]+" "+row[2]
                                val = row[3]
                                if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 
                                    #vaziatMali.append((var,val))

                        if ((j==5) or (j==49) or (j==6) or (j==7) or (j==9) or (j==15) or (j==16) or (j==20) or (j==25) or (j==27) or (j==28) or (j==36) or (j==43) or (j==45) or (j==51)):
                            #if ('۰' not in row):
                                var = row[0].strip()+" "+row[1]
                                val = row[2]
                                #val2 = row[-4] 
                                if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 

                        if ((j==13) or (j==24)):
                            #if ('۰' not in row):
                                    var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]
                                    val = row[4]
                                    if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 
                        
                        if ((j==14) or (j==18) or (j==30) or (j==42) or (j==52)): # or (j==38)
                            #if ('۰' not in row):
                                    var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]
                                    val = row[5]
                                    if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 
                        
                        if ((j==23) or (j==46) or (j==47)):
                            #if ('۰' not in row):
                                    var = row[0].strip()
                                    val = row[1]
                                    if (((var,val) not in row)):
                                        vaziatMali.append((var,val)) 
                    
                    #if (j==48):
                        #if ('۰' not in row):
                               # var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]+" "+row[6]
                              #  val = row[7]
                              #  if (((var,val) not in row)):
                              #      vaziatMali.append((var,val))
                '''
                
                     

            # صورت سود و زیان
            soodVaZian = []
            linkSoodZian = linkNotif+"&sheetid=1"
            soodZianHTML = get_html(linkSoodZian)
            soodZianRows = soodZianHTML.find('tbody').find_all('tr')
            if ("ناخالص" in soodZianRows[3].text.split()):
                for j in range(1,len(soodZianRows)-1):
                    row = soodZianRows[j].text.split()
                    try:
                        if (("تمام" in row) or (j==23)):
                            #if ('۰' not in row):
                                var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]
                                val = row[5]
                                if (((var,val) not in row)):
                                    soodVaZian.append((var,val))

                        elif((j==1) or ("ناخالص" in row) or (j==6) or (j==7) or (j==8) or (j==9) or (j==13) or (j==14) or (j==18) or (j==20) or (j==21) or (j==30) or (j==31)):
                            #if ('۰' not in row):
                                var = row[0].strip()+" "+row[1]
                                val = row[2]
                                if (((var,val) not in row)):
                                    soodVaZian.append((var,val))

                        elif(("ادارى" in row) or (j==10) or (j==15)or (j==22)):
                            #if ('۰' not in row):    or (j==17) 
                                var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]
                                val = row[6]
                                if (((var,val) not in row)):
                                    soodVaZian.append((var,val))

                        elif(("ارزش" in row) or (j==25)):
                            #if ('۰' not in row):
                                var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]+" "+row[6]
                                val = row[7]
                                if (((var,val) not in row)):
                                    soodVaZian.append((var,val))

                        elif(j==11):
                            #if ('۰' not in row):
                                var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]+" "+row[6]+" "+row[7]
                                val = row[8]
                                if (((var,val) not in row)):
                                    soodVaZian.append((var,val))

                        elif((j==23) or (j==24)):
                            #if ('۰' not in row):
                                var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]
                                val = row[4]
                                if (((var,val) not in row)):
                                    soodVaZian.append((var,val))
                        elif((j==26)):
                            #if ('۰' not in row):
                                var = row[0].strip()
                                val = row[1]
                                if (((var,val) not in row)):
                                    soodVaZian.append((var,val))
                    except:
                        soodVaZian = []

            



            # صورت جریان های نقدی 
            jarianNaqd = []
            linkJarianNaqd = linkNotif+"&sheetid=9"
            jarianNaqdHTML = get_html(linkJarianNaqd)
            jarianNaqdRows = jarianNaqdHTML.find('tbody').find_all('tr')
            
            for j in range(1,len(jarianNaqdRows)):
                 row = jarianNaqdRows[j].text.split()
                 try:
                    if ((j==1) or (j==43)):
                        #if ('۰' not in row):
                            var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]
                            val = row[4]
                            if (((var,val) not in row)):
                                        soodVaZian.append((var,val))

                    elif((j==2) or (j==9) or (j==11) or (j==15) or (j==19) or (j==24) or (j==25) or (j==27) or (j==32) or (j==33) or (j==38) or (j==41) or (j==42) or (j==44)):
                        #if ('۰' not in row):
                            var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]
                            val = row[6]
                            if (((var,val) not in row)):
                                        soodVaZian.append((var,val))

                    elif ((j==3) or (j==17) or (j==18) or (j==21)):
                        #if ('۰' not in row):
                            var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]+ " "+row[6]+" "+row[7]+" "+row[8]
                            val = row[9]
                            if (((var,val) not in row)):
                                        soodVaZian.append((var,val))
                    
                    elif ((j==5) or (j==12) or (j==34)):
                        #if ('۰' not in row):
                            var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]+" "+row[6]+" "+row[7]
                            val = row[8]
                            if (((var,val) not in row)):
                                        soodVaZian.append((var,val))
                    
                    elif((j==6) or (j==8) or (j==10) or (j==13) or (j==14) or (j==16) or (j==20) or (j==26) or (j==31) or (j==35) or (j==36) or (j==37)):
                        #if ('۰' not in row):
                            var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]+" "+row[6]
                            val = row[7]
                            if (((var,val) not in row)):
                                        soodVaZian.append((var,val))

                    elif ((j==7)or (j==22) or (j==40)):
                        #if ('۰' not in row):
                            var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]+" "+row[5]+" "+row[6]+" "+row[7]+" "+row[8]+" "+row[9]
                            val = row[10]
                            if (((var,val) not in row)):
                                        soodVaZian.append((var,val))
                    
                    elif ((j==28) or (j==29) or (j==30) or (j==39)):
                        #if ('۰' not in row):
                            var = row[0].strip()+" "+row[1]+" "+row[2]+" "+row[3]+" "+row[4]
                            val = row[5]
                            if (((var,val) not in row)):
                                        soodVaZian.append((var,val))
                        
                    elif ((j==45)):
                        #if ('۰' not in row):
                            var = row[0].strip()+" "+row[1]
                            val = row[2]
                            if (((var,val) not in row)):
                                        soodVaZian.append((var,val))

                    jarianNaqd.append((var,val))
                 except:
                       jarianNaqd = []
                       



            



            # گزارش تولید و فروش ماهانه
            gozareshTolidvaForoosh = []
            '''
                خط های اونو استخراج میکنه
            '''
                     
            data = {"name":name+" "+date,"sal":year,"dore":dore,"vaziatMali":vaziatMali,"soodVaZian":soodVaZian,
                    "jarianNaqd":jarianNaqd,"gozareshTolidvaForoosh":gozareshTolidvaForoosh}
                
#year-dore-talfiqi-date.xlsx

            if (vaziatMali!=[]):
                sheet = excell.active
                #DG    اخرش خطو پلاس پلاس یادت نره
                #sotoonSherkat = "A"
                #rowExcell = sotoonSherkat+counterRowExcell
                counterRowExcell = 4
                counterVaziatMali = 0
                counterSoodVaZian = 0
                counterJarianNaqd = 0

                # اضافه کردن صورت وضعیت مالی به اکسل

                if (len(vaziatMali) > 0 and counterVaziatMali < len(vaziatMali)):
                    for col in range(1,27):
                        column_letter = chr(64 + col)
                        cell = column_letter+str(counterRowExcell)
                        sheet[cell] = data["vaziatMali"][counterVaziatMali][1]
                        counterVaziatMali += 1
                        #counterRowExcell += 1

                    for col in range(1,16):
                        column_letter = "A"+chr(64 + col)
                        cell = column_letter+str(counterRowExcell)
                        sheet[cell] = data["vaziatMali"][counterVaziatMali][1]
                        counterVaziatMali += 1
                else:
                    if (counterVaziatMali < len(vaziatMali)):
                        for col in range(1,27):
                            column_letter = chr(64 + col)
                            cell = column_letter+str(counterRowExcell)
                            sheet[cell] = None
                            counterVaziatMali += 1
                            #counterRowExcell += 1

                        for col in range(1,16):
                            column_letter = "A"+chr(64 + col)
                            cell = column_letter+str(counterRowExcell)
                            sheet[cell] = None
                            counterVaziatMali += 1

                
                # اضافه کردن صورت سود و زیان به اکسل
                if (len(soodVaZian) > 0 and counterSoodVaZian < len(soodVaZian)):
                    for col in range(17,27):
                        column_letter = "A"+chr(64 + col)
                        cell = column_letter+str(counterRowExcell)
                        sheet[cell] = data["soodVaZian"][counterSoodVaZian][1]
                        counterSoodVaZian += 1

                    for col in range(1,14):
                        column_letter = "B"+chr(64 + col)
                        cell = column_letter+str(counterRowExcell)
                        sheet[cell] = data["soodVaZian"][counterSoodVaZian][1]
                        counterSoodVaZian += 1
                else:
                    if (counterSoodVaZian < len(soodVaZian)):
                        for col in range(17,27):
                            column_letter = "A"+chr(64 + col)
                            cell = column_letter+str(counterRowExcell)
                            sheet[cell] = None
                            counterSoodVaZian += 1

                        for col in range(1,14):
                            column_letter = "B"+chr(64 + col)
                            cell = column_letter+str(counterRowExcell)
                            sheet[cell] = None
                            counterSoodVaZian += 1

                # اضافه کردن صورت جریان های نقدی به اکسل
                if (len(jarianNaqd) > 0 and counterJarianNaqd < len(jarianNaqd)):
                    for col in range(16,27):
                        column_letter = "B"+chr(64 + col)
                        cell = column_letter+str(counterRowExcell)
                        sheet[cell] = data["jarianNaqd"][counterJarianNaqd][1]
                        counterJarianNaqd += 1
                    
                    for col in range(1,27):
                        column_letter = "C"+chr(64 + col)
                        cell = column_letter+str(counterRowExcell)
                        sheet[cell] = data["jarianNaqd"][counterJarianNaqd][1]
                        counterJarianNaqd += 1
                    
                    for col in range(1,6):
                        column_letter = "D"+chr(64 + col)
                        cell = column_letter+str(counterRowExcell)
                        sheet[cell] = data["jarianNaqd"][counterJarianNaqd][1]
                        counterJarianNaqd += 1
                else:
                    if (counterJarianNaqd < len(jarianNaqd)):
                        for col in range(16,27):
                            column_letter = "B"+chr(64 + col)
                            cell = column_letter+str(counterRowExcell)
                            sheet[cell] = None
                            counterJarianNaqd += 1
                        
                        for col in range(1,27):
                            column_letter = "C"+chr(64 + col)
                            cell = column_letter+str(counterRowExcell)
                            sheet[cell] = None
                            counterJarianNaqd += 1
                        
                        for col in range(1,6):
                            column_letter = "D"+chr(64 + col)
                            cell = column_letter+str(counterRowExcell)
                            sheet[cell] = None
                            counterJarianNaqd += 1
                
                


                if (talfiqi):         
                    #excell = openpyxl.load_workbook(f'{year}-{dore}-{talfiqi}.xlsx')
                    excellName = f'{year}-{dore}-{talfiqi}.xlsx'

                else:
                    #excell = openpyxl.load_workbook(f'{year}-{dore}.xlsx')
                    excellName = f'{year}-{dore}.xlsx'
                excell.save(excellName)
            else:
                pass
                 



            

                 


                 

            









'''    from selenium import webdriver
    from selenium.webdriver.common.by import By

    # Set up the webdriver
    driver = webdriver.Chrome()

    # Open the webpage
    url = "https://codal.ir/ReportList.aspx?PageNumber=2"
    driver.get(url)

    # Extract the page title
    title_element = driver.find_element(By.XPATH, "//title")
    page_title = title_element.get_attribute("text")
    print("Page Title:", page_title)

    # find name of company
    names = []
    name = driver.find_elements(By.XPATH, "//td[@class='table__content']")
    print(name)
    name0 = name[1]
    print("\n"+name0.get_attribute("data-heading"))

    # Close the webdriver
    driver.quit()
'''
